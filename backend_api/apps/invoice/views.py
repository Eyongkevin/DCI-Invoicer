import io
from collections import namedtuple
from datetime import datetime
from typing import List, Tuple

import boto3
import openpyxl
from apps.core.abstracts import AbstractViewSet
from apps.core.mixins import FilterByLoggedUserMixin
from apps.core.permissions import UserPermission
from apps.user import models as UserModels
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import get_template
from django.urls import reverse
from rest_framework import status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from weasyprint import HTML

from .models import Invoice
from .permissions import UserIDBelongToUser
from .serializers import InvoiceSerializer

# Create your views here.


class InvoiceViewSet(FilterByLoggedUserMixin, AbstractViewSet):
    http_method_names = ("post", "get", "delete")
    permission_classes = (UserPermission, UserIDBelongToUser)
    serializer_class = InvoiceSerializer
    queryset = Invoice.objects.all()

    @action(methods=["post"], detail=False)
    def upload(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["GET"])
def pdf_generate(request, public_id=None):
    def get_user_info(request):
        user_id = str(request.user.public_id)
        profile = UserModels.Profile.objects.get_object_by_user_id(user_id)
        address = UserModels.Address.objects.get_object_by_user_id(user_id)
        company = UserModels.Company.objects.get_object_by_user_id(user_id)
        assignment = UserModels.Assignment.objects.get_object_by_user_id(user_id)

        UserInfo = namedtuple("UserInfo", "profile address company assignment")
        return UserInfo(profile, address, company, assignment)

    def get_auto_info():
        current_year = datetime.now().year
        invoice_number = "009"
        service_period_start = "01/09/2023"
        service_period_end = "30/09/2023"
        generated_date = datetime.now().date().strftime("%d. %h %Y")

        AutoInfo = namedtuple(
            "AutoInfo",
            "year invoice_number service_period_start service_period_end generated_date",
        )
        return AutoInfo(
            current_year,
            invoice_number,
            service_period_start,
            service_period_end,
            generated_date,
        )

    def get_sheet_info(public_id) -> Tuple[List[List[str]], int]:
        invoice = Invoice.objects.get_object_by_public_id(public_id)
        s3 = boto3.client(
            "s3",
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        )
        s3_data = s3.get_object(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=invoice.invoice_xls.name
        )
        contents = s3_data["Body"].read()
        wb = openpyxl.load_workbook(filename=(io.BytesIO(contents)), data_only=True)
        sheet = wb[wb.sheetnames[0]]

        sheet_info = []
        sum_net = 0
        for row in range(1, sheet.max_row + 1):
            sheet_data = []
            for col in range(1, sheet.max_column + 1):
                if col == 1:
                    if sheet.cell(column=col, row=row).value == "t":
                        sum_net += 20 * 8
                    else:
                        sum_net += 20 * 4
                sheet_data.append(sheet.cell(column=col, row=row).value)

            sheet_info.append(sheet_data)

        return sheet_info, sum_net

    if request.method == "GET":
        user_info = get_user_info(request)
        auto_info = get_auto_info()
        sheet_info, sum_net = get_sheet_info(public_id)
        context = {
            "user_info": user_info,
            "auto_info": auto_info,
            "sheet_info": sheet_info,
            "sum_info": {"sum_net": sum_net, "rate_per_eu": 20},
        }
        template = get_template("invoice.html")
        rendered = template.render(context)
        html = HTML(string=rendered)
        pd_file = html.write_pdf()

        # build response
        response = HttpResponse(pd_file, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="invoice.pdf"'
        return response
