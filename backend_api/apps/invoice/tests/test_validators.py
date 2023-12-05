import mock
import pytest
from apps.invoice.validators import validate_file_extension
from django.core.files import File
from openpyxl.utils.exceptions import InvalidFileException
from rest_framework.exceptions import ValidationError


@mock.patch("apps.invoice.validators.load_workbook")
def test_validate_file_extension_pass(mock_load_workbook):
    mock_load_workbook.return_value = None
    file_mock = mock.MagicMock(spec=File)
    file_mock.name = "invoice.xlsx"

    result = validate_file_extension(file_mock)
    assert result is None


@mock.patch("apps.invoice.validators.load_workbook")
def test_validate_file_extension_failed(mock_load_workbook):
    mock_load_workbook.side_effect = InvalidFileException
    file_mock = mock.MagicMock(spec=File)
    file_mock.name = "invoice.pdf"

    with pytest.raises(ValidationError) as err:
        validate_file_extension(file_mock)

    assert err.value.args[0]["detail"] == "Invalid Uploaded file extension"
    assert err.value.args[0]["code"] == "invalid_file_extension_failed"
